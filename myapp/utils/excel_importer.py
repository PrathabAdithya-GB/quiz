from openpyxl import load_workbook
from django.utils.text import slugify

from myapp.models import Category, Quiz, Question, Choice


# ------------------------------------------------
# STEP 1: READ & PARSE EXCEL
# ------------------------------------------------
def parse_excel(file):
    """
    Reads Excel file and returns list of row dictionaries
    """
    wb = load_workbook(file)
    sheet = wb.active

    headers = [cell.value for cell in sheet[1]]

    required_headers = [
        "Category", "Quiz Title", "Question",
        "A", "B", "C", "D",
        "Correct", "Difficulty", "Marks"
    ]

    for h in required_headers:
        if h not in headers:
            raise ValueError(f"Missing required column: {h}")

    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue

        row_data = dict(zip(headers, row))
        data.append(row_data)

    return data


# ------------------------------------------------
# STEP 2: IMPORT INTO DATABASE
# ------------------------------------------------
def import_parsed_data(parsed_data):
    """
    Creates Category, Quiz, Question & Choices from parsed excel data
    """

    for row in parsed_data:

        # -----------------------------
        # CATEGORY
        # -----------------------------
        category, _ = Category.objects.get_or_create(
            name=row["Category"].strip(),
            defaults={"slug": slugify(row["Category"])}
        )

        # -----------------------------
        # QUIZ
        # -----------------------------
        quiz, _ = Quiz.objects.get_or_create(
            title=row["Quiz Title"].strip(),
            defaults={
                "category": category,
                "description": "Imported via Excel",
                "is_published": True,
                "time_limit": 10,
            }
        )

        # -----------------------------
        # QUESTION
        # -----------------------------
        question = Question.objects.create(
            quiz=quiz,
            text=row["Question"].strip(),
            difficulty=(row.get("Difficulty") or "easy").lower(),
            marks=int(row.get("Marks") or 1),
        )

        # -----------------------------
        # CORRECT ANSWERS (A,B,C or A)
        # -----------------------------
        correct_letters = (
            row["Correct"]
            .upper()
            .replace(" ", "")
            .split(",")
        )

        # -----------------------------
        # OPTIONS
        # -----------------------------
        options = {
            "A": row["A"],
            "B": row["B"],
            "C": row["C"],
            "D": row["D"],
        }

        for letter, text in options.items():
            Choice.objects.create(
                question=question,
                text=str(text).strip(),
                is_correct=letter in correct_letters
            )
